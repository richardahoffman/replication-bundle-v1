#!/usr/bin/env python3
"""
run_all.py — CSV-first utilities for the replication bundle.

Commands:
  --summary            Print row counts and column names for each CSV.
  --validate           Validate schemas, enums, IDs, dates, etc.
  --strict             With --validate, treat cross-file issues (e.g., provenance source_id mismatches) as errors.

Layout (relative to this script):
  ../data/*.csv
  ../dictionaries/*.csv
  ../provenance/sources.(csv|xlsx)   [optional; for cross-refs]

Exit codes:
  0 = success
  1 = validation failed (errors present)
"""

from __future__ import annotations
import argparse, csv, sys, re
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DATA = ROOT / "data"
DICT = ROOT / "dictionaries"
PROV = ROOT / "provenance"

# --- make CLI robust to “smart dashes” and Windows-style /flags ---
BAD_DASHES = "–—−‒‐"  # en dash, em dash, minus, figure dash, hyphen
def _sanitized_argv():
    argv = []
    for a in sys.argv[1:]:
        for ch in BAD_DASHES:
            a = a.replace(ch, "-")
        if a.startswith("/"):                 # allow /validate style
            a = "--" + a[1:]
        argv.append(a)
    if not argv:                              # default behavior: show summary
        argv = ["--summary"]
    return argv

# ---------- CSV helpers ----------

def read_csv(path: Path) -> list[dict]:
    """BOM-safe CSV reader that trims header/value whitespace."""
    rows: list[dict] = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            try:
                raw_headers = next(reader)
            except StopIteration:
                return []  # empty file
            headers = [(h or "").replace("\ufeff", "").strip() for h in raw_headers]
            for raw in reader:
                row = {}
                for i, h in enumerate(headers):
                    if i < len(raw):
                        v = raw[i]
                        row[h] = v.strip() if isinstance(v, str) else v
                    else:
                        row[h] = ""
                rows.append(row)
    except FileNotFoundError:
        return []
    return rows

def require_columns(rows: list[dict], required: list[str]) -> list[str]:
    errors: list[str] = []
    if not rows:
        errors.append("file has no rows (after header) or is empty")
        return errors
    headers = list(rows[0].keys())
    miss = [c for c in required if c not in headers]
    if miss:
        errors.append(f"missing required columns: {', '.join(miss)}")
    return errors

def ok_iso_date(value: str) -> bool:
    """Accepts YYYY, YYYY-MM, YYYY-MM-DD with calendar validation."""
    if not value or not isinstance(value, str):
        return False
    m = re.fullmatch(r"(\d{4})(?:-(\d{2})(?:-(\d{2}))?)?", value.strip())
    if not m:
        return False
    y = int(m.group(1))
    if y < 1 or y > 9999:
        return False
    mm = m.group(2)
    dd = m.group(3)
    try:
        if dd:
            datetime(y, int(mm), int(dd))
        elif mm:
            datetime(y, int(mm), 1)
        else:
            pass
        return True
    except ValueError:
        return False

def unique(values: list[str]) -> bool:
    seen = [v for v in values if v not in ("", None)]
    return len(set(seen)) == len(seen)

def load_provenance_ids() -> set[str]:
    """Collect source_id keys from provenance/sources.csv or sources.xlsx (optional)."""
    ids: set[str] = set()
    csv_path = PROV / "sources.csv"
    xlsx_path = PROV / "sources.xlsx"

    def add_csv(p: Path):
        rows = read_csv(p)
        if not rows:
            return
        header_keys = rows[0].keys()
        key = "source_id" if "source_id" in header_keys else ("id" if "id" in header_keys else None)
        if key:
            for r in rows:
                v = (r.get(key) or "").strip()
                if v:
                    ids.add(v)

    if csv_path.exists():
        add_csv(csv_path)
    elif xlsx_path.exists():
        try:
            import openpyxl  # optional dependency
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = None
            for cand in ("source_id", "id"):
                if cand in header:
                    idx = header.index(cand)
                    break
            if idx is not None:
                for row in ws.iter_rows(min_row=2):
                    v = row[idx].value
                    if v is not None:
                        ids.add(str(v).strip())
        except Exception:
            pass
    return ids

def warn_unknown_values(rows: list[dict], col: str, allowed: set[str]) -> list[str]:
    warnings = []
    seen = set()
    for r in rows:
        v = (r.get(col) or "").strip()
        if v and v not in allowed:
            seen.add(v)
    if seen:
        warnings.append(f"{col}: found {len(seen)} value(s) not in recommended set: {sorted(seen)}")
    return warnings

# ---------- Schemas ----------

SCHEMAS: dict[str, dict] = {
    "table1_promises_vs_outcomes.csv": {
        "required": ["section", "promise", "administration", "reality", "sources"],
        "id_col": "t1_id",
        "id_regex": r"^T1-\d{3}$",
    },
    "appendix_b_evidence_to_claim.csv": {
        "required": ["domain", "claim_id", "claim_text", "test_type", "method", "sources", "notes"],
        "id_col": "claim_id",
        "id_regex": r"^B-\d{3}$",
        "enums": {
            "test_type": {"straw_in_the_wind", "hoop", "smoking_gun", "doubly_decisive"},
        },
    },
    "appendix_c_indicators.csv": {
        "required": ["indicator", "definition", "unit", "source_id", "frequency", "notes"],
        "id_col": "indicator_id",
        "id_regex": r"^C-\d{3}$",
        "enums": {
            "frequency": {"annual", "quarterly", "monthly", "weekly", "daily", "ad hoc"},
        },
    },
    "appendix_e_timelines.csv": {
        "required": ["domain", "date", "event", "source_id"],
        "id_col": "event_id",
        "id_regex": r"^E-\d{3}$",
        "enums": {
            "date_precision": {"day", "month", "year"},
        },
        "date_col": "date",
    },
    "abbreviations.csv": {
        "required": ["term", "expansion", "scope"],
        "recommended_scopes": {
            "Law", "Agency", "Oversight", "Index/Metric", "NGO", "Think Tank/NGO",
            "Intl. Org", "Non-state actor", "Policy", "Program", "Law/Authority",
            "Company", "Court", "General", "Tech", "Admin/NEPA", "Agency/Grants",
                          "Budget/Accounting", "Law/Program", "NGO/Media", "Presidential Action"
        },
        "unique_col": "term",
    },
}

# ---------- Validators ----------

def validate_file(name: str, strict: bool, prov_ids: set[str]) -> tuple[list[str], list[str]]:
    path = DATA / name
    errors: list[str] = []
    warnings: list[str] = []

    if not path.exists():
        return [f"missing file: {path}"], warnings

    rows = read_csv(path)
    if not rows:
        return ["file has no rows (or is empty)"], warnings

    schema = SCHEMAS[name]

    errors += require_columns(rows, schema["required"])
    if errors:
        return errors, warnings

    headers = rows[0].keys()
    if "" in headers:
        errors.append("blank column name in header")

    # Optional: id column & uniqueness
    id_col = schema.get("id_col")
    id_re = schema.get("id_regex")
    if id_col and id_col in headers:
        ids = [(r.get(id_col) or "").strip() for r in rows]
        bad = [v for v in ids if v and (not re.fullmatch(id_re, v))]
        if bad:
            errors.append(f"{id_col} format violations (expected {id_re}): {sorted(set(bad))[:10]}{'...' if len(set(bad))>10 else ''}")
        if not unique(ids):
            dupes = [k for k, c in Counter([v for v in ids if v]).items() if c > 1]
            errors.append(f"{id_col} duplicate values: {sorted(dupes)[:10]}{'...' if len(dupes)>10 else ''}")

    # Enums (hard errors)
    for col, allowed in schema.get("enums", {}).items():
        if col in headers:
            invalid = sorted({(r.get(col) or "").strip() for r in rows if r.get(col) and (r.get(col).strip() not in allowed)})
            if invalid:
                errors.append(f"{col}: {len(invalid)} value(s) not in {sorted(allowed)} — {invalid[:10]}{'...' if len(invalid)>10 else ''}")

    # Date validation
    date_col = schema.get("date_col")
    if date_col:
        bad_dates = sorted({(r.get(date_col) or "").strip() for r in rows if r.get(date_col) and not ok_iso_date(r.get(date_col))})
        if bad_dates:
            errors.append(f"{date_col}: invalid ISO date(s) (YYYY[-MM[-DD]]): {bad_dates[:10]}{'...' if len(bad_dates)>10 else ''}")

    # Recommended values (warnings)
    if name == "abbreviations.csv":
        if "scope" in headers:
            warnings += warn_unknown_values(rows, "scope", SCHEMAS[name]["recommended_scopes"])
        ucol = SCHEMAS[name].get("unique_col")
        if ucol and ucol in headers:
            terms = [ (r.get(ucol) or "").strip() for r in rows if (r.get(ucol) or "").strip() ]
            dupes = [k for k, c in Counter(terms).items() if c > 1]
            if dupes:
                warnings.append(f"{ucol}: duplicate values: {sorted(dupes)[:10]}{'...' if len(dupes)>10 else ''}")

    # Cross-file: provenance source_id presence
    if name in ("appendix_c_indicators.csv", "appendix_e_timelines.csv") and prov_ids:
        col = "source_id"
        if col in headers:
            missing = sorted({(r.get(col) or "").strip() for r in rows
                              if (r.get(col) or "").strip() and (r.get(col).strip() not in prov_ids)})
            if missing:
                msg = f"{col}: {len(missing)} value(s) not found in provenance sources: {missing[:10]}{'...' if len(missing)>10 else ''}"
                (errors if strict else warnings).append(msg)

    # CSV hygiene (warnings)
    trailing = []
    for i, r in enumerate(rows, start=2):
        for k, v in r.items():
            if isinstance(v, str) and v != v.strip():
                trailing.append((i, k))
                if len(trailing) >= 10:
                    break
        if len(trailing) >= 10:
            break
    if trailing:
        warnings.append(f"trailing whitespace in {len(trailing)} cell(s) (first 10 shown): {trailing}")

    return errors, warnings

def validate_all(strict: bool=False) -> int:
    prov_ids = load_provenance_ids()
    all_errors: dict[str, list[str]] = defaultdict(list)
    all_warnings: dict[str, list[str]] = defaultdict(list)
    exit_code = 0

    for name in SCHEMAS.keys():
        errs, warns = validate_file(name, strict=strict, prov_ids=prov_ids)
        if errs:
            exit_code = 1
            all_errors[name].extend(errs)
        if warns:
            all_warnings[name].extend(warns)

    print("\nVALIDATION REPORT\n==================")
    for name in SCHEMAS.keys():
        print(f"\n{name}")
        if all_errors[name]:
            for e in all_errors[name]:
                print(f"  ERROR   - {e}")
        else:
            print("  OK")
        for w in all_warnings[name]:
            print(f"  warning - {w}")

    if exit_code == 0:
        print("\nAll required checks passed.")
    else:
        print("\nValidation failed. See errors above.")
    return exit_code

def summary():
    print("\nDATA SUMMARY\n============")
    for name in SCHEMAS.keys():
        p = DATA / name
        if not p.exists():
            print(f"{name:35} MISSING")
            continue
        rows = read_csv(p)
        if not rows:
            print(f"{name:35}    0 rows | (empty)")
            continue
        cols = list(rows[0].keys())
        print(f"{name:35} {len(rows):6d} rows | {len(cols):2d} cols | {', '.join(cols)}")

def main():
    ap = argparse.ArgumentParser(description="Replication bundle utilities (CSV-first).")
    ap.add_argument("--summary", action="store_true", help="print row counts and column lists")
    ap.add_argument("--validate", action="store_true", help="run schema/consistency checks")
    ap.add_argument("--strict", action="store_true", help="treat cross-file warnings (e.g., provenance) as errors")

    args = ap.parse_args(_sanitized_argv())   # use sanitized argv

    ran = False
    if args.summary:
        summary(); ran = True
    if args.validate:
        code = validate_all(strict=args.strict); ran = True; sys.exit(code)
    if not ran:
        ap.print_help()

if __name__ == "__main__":
    main()
